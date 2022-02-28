import telebot
import openpyxl
from docxtpl import DocxTemplate
import docx
from telebot import types
from collections import defaultdict


token = '5005977482:AAHTnugVRwfrbV_Mn7YV5qbieWs4POmqxTg'
bot = telebot.TeleBot(token, skip_pending=True)


schetchik_1 = {}  # Счетчик основных вопросов
schetchik_2 = {}  # Счетчик дополнительных вопросов
schetchik = {}    # Счетчик всех вопросов
number_dict = {}
date_time_dict = {}
name_dict = {}
place_dict = {}
condition_dict = {}
medical_equipment_dict = {}
fall_dict = {}
unauthorized_leave_dict = {}
medical_drug_dict = {}
other_dict = {}
description_dict = {}
consequence_dict = {}
measures_dict = {}
note_dict = {}
list_for_comment = defaultdict(list)


wb_a = openpyxl.reader.excel.load_workbook(filename="Ambulator.xlsx")
wb_a.active = 0
sheet_a = wb_a.active


wb_s = openpyxl.reader.excel.load_workbook(filename="Stationar.xlsx")
wb_s.active = 0
sheet_s = wb_s.active


def table(message_for_table, table_element, next_table_element, text):
    if message_for_table.text == "Начать заполнение заново":
        mesg = bot.send_message(message_for_table.chat.id, 'Вы начали заполнение заново. Напишите боту повторную команду "/start"')
        bot.register_next_step_handler(mesg, welcome)
    else:
        table_element[message_for_table.chat.id] = message_for_table.text
        mesg = bot.send_message(message_for_table.chat.id, text)
        bot.register_next_step_handler(mesg, next_table_element)
        print(table_element)


def clear_dict():
    number_dict.clear()
    date_time_dict.clear()
    name_dict.clear()
    place_dict.clear()
    condition_dict.clear()
    medical_equipment_dict.clear()
    fall_dict.clear()
    unauthorized_leave_dict.clear()
    medical_drug_dict.clear()
    other_dict.clear()
    description_dict.clear()
    consequence_dict.clear()
    measures_dict.clear()
    note_dict.clear()


def create_keyboard(call_create):
    global schetchik
    global sheet_a
    schetchik[call_create.chat.id] += 1
    # --------------------------------------Для вопросов 1 клавиуатура---------------------------------------
    key_for_test_1 = types.InlineKeyboardMarkup()
    btn_yes_1 = types.InlineKeyboardButton('Да', callback_data='yes_1')
    btn_no_1 = types.InlineKeyboardButton('Нет', callback_data='no_1')
    btn_act_1 = types.InlineKeyboardButton('Нормативный акт', callback_data='act_1')
    btn_comment_1 = types.InlineKeyboardButton('Комментарии', callback_data='comment_1')
    key_for_test_1.row(btn_yes_1, btn_no_1)
    key_for_test_1.row(btn_act_1, btn_comment_1)
    # --------------------------------------Для вопросов 2 клавиатура----------------------------------------
    key_for_test_2 = types.InlineKeyboardMarkup()
    btn_yes_2 = types.InlineKeyboardButton('Да', callback_data='yes_2')
    btn_no_2 = types.InlineKeyboardButton('Нет', callback_data='no_2')
    btn_act_2 = types.InlineKeyboardButton('Нормативный акт', callback_data='act_2')
    btn_comment_2 = types.InlineKeyboardButton('Комментарии', callback_data='comment_2')
    key_for_test_2.row(btn_yes_2, btn_no_2)
    key_for_test_2.row(btn_act_2, btn_comment_2)
    # --------------------------------------Для вопросов 3 клавиатура----------------------------------------
    key_for_test_3 = types.InlineKeyboardMarkup()
    btn_yes_3 = types.InlineKeyboardButton('Да', callback_data='yes_3')
    btn_no_3 = types.InlineKeyboardButton('Нет', callback_data='no_3')
    btn_act_3 = types.InlineKeyboardButton('Нормативный акт', callback_data='act_3')
    btn_comment_3 = types.InlineKeyboardButton('Комментарии', callback_data='comment_3')
    key_for_test_3.row(btn_yes_3, btn_no_3)
    key_for_test_3.row(btn_act_3, btn_comment_3)
    if schetchik[call_create.chat.id] == 193:
        if schetchik_1[call_create.chat.id] == 133 and schetchik_2[call_create.chat.id] >= 18:
            bot.send_message(call_create.chat.id, 'Медицинская помощь оказана в соответствии с порядками, стандартами и клиническими рекомендациями')
            bot.send_message(call_create.chat.id, 'Ваши комментарии:')
            bot.send_message(call_create.chat.id, str(list_for_comment[call_create.chat.id]).strip('[]'))
        else:
            bot.send_message(call_create.chat.id, 'Медицинская помощь оказана с дефектами')
            bot.send_message(call_create.chat.id, 'Ваши комментарии:')
            bot.send_message(call_create.chat.id, str(list_for_comment[call_create.chat.id]).strip('[]'))
    if sheet_a['B' + str(schetchik[call_create.chat.id])].value == 1:  # Проверка вопроса на то, что он основной
        print('Вопрос значения 1', schetchik[call_create.chat.id], schetchik_1[call_create.chat.id], schetchik_2[call_create.chat.id])
        bot.send_message(call_create.chat.id, 'Вопрос №' + str(schetchik[call_create.chat.id]))
        bot.send_message(call_create.chat.id, sheet_a['A' + str(schetchik[call_create.chat.id])].value, reply_markup=key_for_test_1)
    if sheet_a['B' + str(schetchik[call_create.chat.id])].value == 2:  # Проверка вопроса на то, что он дополнительный
        print('Вопрос значения 2', schetchik[call_create.chat.id], schetchik_1[call_create.chat.id], schetchik_2[call_create.chat.id])
        bot.send_message(call_create.chat.id, 'Вопрос №' + str(schetchik[call_create.chat.id]))
        bot.send_message(call_create.chat.id, sheet_a['A' + str(schetchik[call_create.chat.id])].value, reply_markup=key_for_test_2)
    if sheet_a['B' + str(schetchik[call_create.chat.id])].value == 3:  # Проверка вопроса на то, что он необязательный
        print('Вопрос значения 3', schetchik[call_create.chat.id], schetchik_1[call_create.chat.id], schetchik_2[call_create.chat.id])
        bot.send_message(call_create.chat.id, 'Вопрос №' + str(schetchik[call_create.chat.id]))
        bot.send_message(call_create.chat.id, sheet_a['A' + str(schetchik[call_create.chat.id])].value, reply_markup=key_for_test_3)


def create_keyboard_st(call_create):
    global schetchik
    global sheet_s
    global list_for_comment
    schetchik[call_create.chat.id] += 1
    # --------------------------------------Для вопросов 1 клавиуатура---------------------------------------
    key_for_test_1_st = types.InlineKeyboardMarkup()
    btn_yes_1_st = types.InlineKeyboardButton('Да', callback_data='yes_1_st')
    btn_no_1_st = types.InlineKeyboardButton('Нет', callback_data='no_1_st')
    btn_act_1_st = types.InlineKeyboardButton('Нормативный акт', callback_data='act_1_st')
    btn_comment_1_st = types.InlineKeyboardButton('Комментарии', callback_data='comment_1_st')
    key_for_test_1_st.row(btn_yes_1_st, btn_no_1_st)
    key_for_test_1_st.row(btn_act_1_st, btn_comment_1_st)
    # --------------------------------------Для вопросов 2 клавиатура----------------------------------------
    key_for_test_2_st = types.InlineKeyboardMarkup()
    btn_yes_2_st = types.InlineKeyboardButton('Да', callback_data='yes_2_st')
    btn_no_2_st = types.InlineKeyboardButton('Нет', callback_data='no_2_st')
    btn_act_2_st = types.InlineKeyboardButton('Нормативный акт', callback_data='act_2_st')
    btn_comment_2_st = types.InlineKeyboardButton('Комментарии', callback_data='comment_2_st')
    key_for_test_2_st.row(btn_yes_2_st, btn_no_2_st)
    key_for_test_2_st.row(btn_act_2_st, btn_comment_2_st)
    # --------------------------------------Для вопросов 3 клавиатура----------------------------------------
    key_for_test_3_st = types.InlineKeyboardMarkup()
    btn_yes_3_st = types.InlineKeyboardButton('Да', callback_data='yes_3_st')
    btn_no_3_st = types.InlineKeyboardButton('Нет', callback_data='no_3_st')
    btn_act_3_st = types.InlineKeyboardButton('Нормативный акт', callback_data='act_3_st')
    btn_comment_3_st = types.InlineKeyboardButton('Комментарии', callback_data='comment_3_st')
    key_for_test_3_st.row(btn_yes_3_st, btn_no_3_st)
    key_for_test_3_st.row(btn_act_3_st, btn_comment_3_st)
    if schetchik[call_create.chat.id] == 200:
        if schetchik_1[call_create.chat.id] == 157 and schetchik_2[call_create.chat.id] >= 15:
            bot.send_message(call_create.chat.id, 'Медицинская помощь оказана в соответствии с порядками, стандартами и клиническими рекомендациями')
            bot.send_message(call_create.chat.id, 'Ваши комментарии:')
            bot.send_message(call_create.chat.id, str(list_for_comment[call_create.chat.id]).strip('[]'))
        else:
            bot.send_message(call_create.chat.id, 'Медицинская помощь оказана с дефектами')
            bot.send_message(call_create.chat.id, 'Ваши комментарии:')
            bot.send_message(call_create.chat.id, str(list_for_comment[call_create.chat.id]).strip('[]'))
    if sheet_s['B' + str(schetchik[call_create.chat.id])].value == 1:  # Проверка вопроса на то, что он основной
        print('Вопрос значения 1', schetchik[call_create.chat.id], schetchik_1[call_create.chat.id], schetchik_2[call_create.chat.id])
        bot.send_message(call_create.chat.id, 'Вопрос №' + str(schetchik[call_create.chat.id]))
        bot.send_message(call_create.chat.id, sheet_s['A' + str(schetchik[call_create.chat.id])].value, reply_markup=key_for_test_1_st)
    if sheet_s['B' + str(schetchik[call_create.chat.id])].value == 2:  # Проверка вопроса на то, что он дополнительный
        print('Вопрос значения 2', schetchik[call_create.chat.id], schetchik_1[call_create.chat.id], schetchik_2[call_create.chat.id])
        bot.send_message(call_create.chat.id, 'Вопрос №' + str(schetchik[call_create.chat.id]))
        bot.send_message(call_create.chat.id, sheet_s['A' + str(schetchik[call_create.chat.id])].value, reply_markup=key_for_test_2_st)
    if sheet_s['B' + str(schetchik[call_create.chat.id])].value == 3:  # Проверка вопроса на то, что он необязательный
        print('Вопрос значения 3', schetchik[call_create.chat.id], schetchik_1[call_create.chat.id], schetchik_2[call_create.chat.id])
        bot.send_message(call_create.chat.id, 'Вопрос №' + str(schetchik[call_create.chat.id]))
        bot.send_message(call_create.chat.id, sheet_s['A' + str(schetchik[call_create.chat.id])].value, reply_markup=key_for_test_3_st)


@bot.message_handler(commands=['start'])
def welcome(message):
    global schetchik, schetchik_1, schetchik_2
    schetchik[message.chat.id] = 3  # Сщетчик в исходное значение
    schetchik_1[message.chat.id] = 0
    schetchik_2[message.chat.id] = 0
    #----------------------------------------------ОЧИЩАЕМ СЛОВАРИ------------------------------------------------------
    #--------------------------------------------------KEYBOARD---------------------------------------------------------
    print(name_dict, number_dict, date_time_dict)
    markup_main = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True, row_width=2)
    button1 = types.KeyboardButton("Начать заполнение заново")
    button2 = types.KeyboardButton("/start")
    markup_main.add(button1)
    markup_main.add(button2)

    markup2 = types.InlineKeyboardMarkup(row_width=2)
    item3 = types.InlineKeyboardButton("Пройти тест", callback_data='test')
    item4 = types.InlineKeyboardButton("Заполнить таблицу", callback_data='sheet')
    markup2.add(item3, item4)

    bot.send_message(message.chat.id, 'Привет', reply_markup=markup_main)
    bot.send_message(message.chat.id, 'Выберите, что Вы хотите пройти:', reply_markup=markup2)


@bot.message_handler(content_types=['text'])
def lalala(message):
    if message.text == "Начать заполнение заново":
        mesg = bot.send_message(message.chat.id, 'Вы начали заполнение заново. Напишите боту повторную команду "/start"')
        bot.register_next_step_handler(mesg, welcome)


#_________________________________АМБУЛАТОРНЫЕ УСЛОВИЯ________________________________________
def number_ambulat(message):
    message_for_table = message
    table(message_for_table, number_dict, date_ambulat, 'Дата и время события')


def date_ambulat(message):
    message_for_table = message
    table(message_for_table, date_time_dict, name_ambulat, 'ФИО пострадавшего')


def name_ambulat(message):
    message_for_table = message
    table(message_for_table, name_dict, place_ambulat, 'Место происшествия нежелательного события')


def place_ambulat(message):
    message_for_table = message
    table(message_for_table, place_dict, condition_ambulat, 'Неотложное состояние')


def condition_ambulat(message):
    message_for_table = message
    table(message_for_table, condition_dict, medical_equipment_ambulat, 'Событие, связанное с медицинским оборудованием')


def medical_equipment_ambulat(message):
    message_for_table = message
    table(message_for_table, medical_equipment_dict, fall_ambulat, 'Падение на территории')


def fall_ambulat(message):
    message_for_table = message
    table(message_for_table, fall_dict, unauthorized_leave_ambulat, 'Побег с территории медицинской организации')


def unauthorized_leave_ambulat(message):
    message_for_table = message
    table(message_for_table, unauthorized_leave_dict, medical_drug_ambulat, 'Событие, связано с лекарственным средством')


def medical_drug_ambulat(message):
    message_for_table = message
    table(message_for_table, medical_drug_dict, other_ambulat, 'Другое нежелательное событие')


def other_ambulat(message):
    message_for_table = message
    table(message_for_table, other_dict, description_ambulat, 'Описание обстоятельств, при которых произошло нежелательное событие')


def description_ambulat(message):
    message_for_table = message
    table(message_for_table, description_dict, consequence_ambulat, 'Последствия нежелательного события')


def consequence_ambulat(message):
    message_for_table = message
    table(message_for_table, consequence_dict, measures_ambulat, 'Принятые меры по устранению последствий нежелательного события')


def measures_ambulat(message):
    message_for_table = message
    table(message_for_table, measures_dict, note_ambulat, 'Примечание')


#-----------------------------------------------------RUN TO EXCEL AND WORD---------------------------------------------
def note_ambulat(message):
    if message.text == "Начать заполнение заново":
        mesg = bot.send_message(message.chat.id, 'Вы начали заполнение заново. Напишите боту повторную команду "/start"')
        bot.register_next_step_handler(mesg, welcome)
    else:
        note_dict[message.chat.id] = message.text
        wb = openpyxl.reader.excel.load_workbook(filename="unwanted_ambulat.xlsx")
        ws = wb.active
        #-----------------ПРИСВИВАЕМ ЗНАЧЕНИЯ СЛОВАРЕЙ ОПРЕДЕЛЕННОГО ПОЛЬЗОВАТЕЛЯ ПЕРЕМЕННОЙ ДЛЯ ЗАПИСИ----------------
        number = number_dict[message.chat.id]
        date_time = date_time_dict[message.chat.id]
        name = name_dict[message.chat.id]
        place = place_dict[message.chat.id]
        condition = condition_dict[message.chat.id]
        medical_equipment = medical_equipment_dict[message.chat.id]
        fall = fall_dict[message.chat.id]
        unauthorized_leave = unauthorized_leave_dict[message.chat.id]
        medical_drug = medical_drug_dict[message.chat.id]
        other = other_dict[message.chat.id]
        description = description_dict[message.chat.id]
        consequence = consequence_dict[message.chat.id]
        measures = measures_dict[message.chat.id]
        note = note_dict[message.chat.id]
        #-----------------------------------------ЗАПИСЬ ПЕРЕМЕННЫХ В ФАЙЛ EXCEL---------------------------------------
        data = [(number, date_time, name, place, condition, medical_equipment, fall, unauthorized_leave, medical_drug,
                 other, description, consequence, measures, note)]
        for i in range(1, 100):
            if ws['A' + str(i)].value == None:
                for j in data:
                    ws.append(j)
                break

        wb.save("unwanted_ambulat.xlsx")
        #---------------------------------------ЗАПИСЬ ИНФОРМАЦИИ В ФАЙЛ WORD------------------------------------------
        template = DocxTemplate('template_magazine.docx')
        data = []
        for i in range(4, ws.max_row + 1):
            data.append({
                'number': ws.cell(i, 1).value,
                'date_time': ws.cell(i, 2).value,
                'name': ws.cell(i, 3).value,
                'place': ws.cell(i, 4).value,
                'condition': ws.cell(i, 5).value,
                'medical_equipment': ws.cell(i, 6).value,
                'fall': ws.cell(i, 7).value,
                'unauthorized_leave': ws.cell(i, 8).value,
                'medical_drug': ws.cell(i, 9).value,
                'other': ws.cell(i, 10).value,
                'description': ws.cell(i, 11).value,
                'consequence': ws.cell(i, 12).value,
                'measures': ws.cell(i, 13).value,
                'note': ws.cell(i, 14).value,
            })

        context = {
            'data': data,
        }

        template.render(context)
        template.save('ready_made_magazine_ambulat.docx')
        doc = open('ready_made_magazine_ambulat.docx', 'rb')
        bot.send_document(message.chat.id, doc)


#---------------------------------------------ЗАПОЛНЕНИЕ ТАБЛИЦЫ-------------------------------------------------------
#--------------------------------------------СТАЦИОНАРНЫЕ УСЛОВИЯ------------------------------------------------------


def number_station(message):
    message_for_table = message
    table(message_for_table, number_dict, date_station, 'Дата и время события')


def date_station(message):
    message_for_table = message
    table(message_for_table, date_time_dict, name_station, 'ФИО пострадавшего')


def name_station(message):
    message_for_table = message
    table(message_for_table, name_dict, place_station, 'Место происшествия нежелательного события')


def place_station(message):
    message_for_table = message
    table(message_for_table, place_dict, condition_station, 'Неотложное состояние')


def condition_station(message):
    message_for_table = message
    table(message_for_table, condition_dict, medical_equipment_station, 'Событие, связанное с медицинским оборудованием')


def medical_equipment_station(message):
    message_for_table = message
    table(message_for_table, medical_equipment_dict, fall_station, 'Падение на территории')


def fall_station(message):
    message_for_table = message
    table(message_for_table, fall_dict, unauthorized_leave_station, 'Побег с территории медицинской организации')


def unauthorized_leave_station(message):
    message_for_table = message
    table(message_for_table, unauthorized_leave_dict, medical_drug_station, 'Событие, связано с лекарственным средством')


def medical_drug_station(message):
    message_for_table = message
    table(message_for_table, medical_drug_dict, other_station, 'Другое нежелательное событие')


def other_station(message):
    message_for_table = message
    table(message_for_table, other_dict, description_station, 'Описание обстоятельств, при которых произошло нежелательное событие')


def description_station(message):
    message_for_table = message
    table(message_for_table, description_dict, consequence_station, 'Последствия нежелательного события')


def consequence_station(message):
    message_for_table = message
    table(message_for_table, consequence_dict, measures_station, 'Принятые меры по устранению последствий нежелательного события')


def measures_station(message):
    message_for_table = message
    table(message_for_table, measures_dict, note_station, 'Примечание')


#----------------------------------------------ЗАПИСЬ ПЕРЕМЕННЫХ В ФАЙЛ EXCEL------------------------------------------
def note_station(message):
    if message.text == "Начать заполнение заново":
        mesg = bot.send_message(message.chat.id, 'Вы начали заполнение заново. Напишите боту повторную команду "/start"')
        bot.register_next_step_handler(mesg, welcome)
    else:
        note_dict[message.chat.id] = message.text
        wb = openpyxl.reader.excel.load_workbook(filename="unwanted_station.xlsx")
        ws = wb.active
        number = number_dict[message.chat.id]
        date_time = date_time_dict[message.chat.id]
        name = name_dict[message.chat.id]
        place = place_dict[message.chat.id]
        condition = condition_dict[message.chat.id]
        medical_equipment = medical_equipment_dict[message.chat.id]
        fall = fall_dict[message.chat.id]
        unauthorized_leave = unauthorized_leave_dict[message.chat.id]
        medical_drug = medical_drug_dict[message.chat.id]
        other = other_dict[message.chat.id]
        description = description_dict[message.chat.id]
        consequence = consequence_dict[message.chat.id]
        measures = measures_dict[message.chat.id]
        note = note_dict[message.chat.id]
        data = [(number, date_time, name, place, condition, medical_equipment, fall, unauthorized_leave, medical_drug,
                 other, description, consequence, measures, note)]
        for i in range(1, 100):
            if ws['A' + str(i)].value == None:
                for j in data:
                    ws.append(j)
                break

        wb.save("unwanted_station.xlsx")
        #------------------------------------------ЗАПИСЬ ПЕРЕМЕННЫХ В ФАЙЛ WORD----------------------------------------
        template = DocxTemplate('template_magazine.docx')
        data = []
        for i in range(4, ws.max_row + 1):
            data.append({
                'number': ws.cell(i, 1).value,
                'date_time': ws.cell(i, 2).value,
                'name': ws.cell(i, 3).value,
                'place': ws.cell(i, 4).value,
                'condition': ws.cell(i, 5).value,
                'medical_equipment': ws.cell(i, 6).value,
                'fall': ws.cell(i, 7).value,
                'unauthorized_leave': ws.cell(i, 8).value,
                'medical_drug': ws.cell(i, 9).value,
                'other': ws.cell(i, 10).value,
                'description': ws.cell(i, 11).value,
                'consequence': ws.cell(i, 12).value,
                'measures': ws.cell(i, 13).value,
                'note': ws.cell(i, 14).value,
            })

        context = {
            'data': data
        }

        template.render(context)
        template.save('ready_made_magazine_station.docx')
        doc = open('ready_made_magazine_station.docx', 'rb')
        bot.send_document(message.chat.id, doc)


def comment(message):
    global list_for_comment
    a = message.chat.id
    b = message.text
    list_for_comment[a].append(b)
    print(list_for_comment)
    bot.send_message(message.chat.id, 'Я запомнил Ваш комментарий')


#@bot.message_handler(func=lambda c:True, content_types=['text'])#этот блок выполнится если юзер отправит боту сообщение
#def info_message(message):
#   bot.edit_message_reply_markup(message.chat.id, message_id=message.message_id-1, reply_markup='')


@bot.callback_query_handler(func=lambda call: True)
def callback_inline(call):
    try:
        if call.message:
            if call.data == 'test':
                bot.send_message(call.message.chat.id, 'Вы выбрали тест')
                markup_for_test = types.InlineKeyboardMarkup(row_width=2)
                item1_for_test = types.InlineKeyboardButton("амбулаторные условия", callback_data='ambulat_for_test')
                item2_for_test = types.InlineKeyboardButton("стационарные условия", callback_data='station_for_test')
                markup_for_test.add(item1_for_test, item2_for_test)
                bot.send_message(call.message.chat.id, 'Вы выбрали тест, теперь выберите условия оказания медицинской помощи:', reply_markup=markup_for_test)
            elif call.data == 'sheet':
                bot.send_message(call.message.chat.id, 'Вы выбрали таблицу', reply_markup=None)
                markup_for_sheet = types.InlineKeyboardMarkup(row_width=2)
                item1_for_sheet = types.InlineKeyboardButton("амбулаторные условия", callback_data='ambulat_for_sheet')
                item2_for_sheet = types.InlineKeyboardButton("стационарные условия", callback_data='station_for_sheet')
                markup_for_sheet.add(item1_for_sheet, item2_for_sheet)
                bot.send_message(call.message.chat.id, 'Вы выбрали таблицу, теперь выберите условия оказания медицинской помощи:', reply_markup=markup_for_sheet)
            elif call.data == 'ambulat_for_sheet':
                mesg = bot.send_message(call.message.chat.id, '№ п/п')
                bot.register_next_step_handler(mesg, number_ambulat)
            elif call.data == 'station_for_sheet':
                mesg = bot.send_message(call.message.chat.id, '№ п/п')
                bot.register_next_step_handler(mesg, number_station)
#----------------------------------------------НАЧИНАЮТСЯ АМБУЛАТОРНЫЕ УСЛОВИЯ------------------------------------------
            elif call.data == 'ambulat_for_test':
                global schetchik
                global schetchik_1
                global schetchik_2
                print(call.message.chat.id)
                schetchik[call.message.chat.id] += 1
                #schetchik += 1
                key_for_test_1 = types.InlineKeyboardMarkup()
                btn_yes_1 = types.InlineKeyboardButton('Да', callback_data='yes_1')
                btn_no_1 = types.InlineKeyboardButton('Нет', callback_data='no_1')
                btn_act_1 = types.InlineKeyboardButton('Нормативный акт', callback_data='act_1')
                btn_comment_1 = types.InlineKeyboardButton('Комментарии', callback_data='comment_1')
                key_for_test_1.row(btn_yes_1, btn_no_1)
                key_for_test_1.row(btn_act_1, btn_comment_1)
                wb = openpyxl.reader.excel.load_workbook(filename="Ambulator.xlsx")
                wb.active = 0
                print("Программа зашла в амбулат фор тест который с клавиатуры")
                sheet = wb.active
                bot.send_message(call.message.chat.id, 'Вопрос №' + str(schetchik[call.message.chat.id]))
                bot.send_message(call.message.chat.id, sheet['A' + str(schetchik[call.message.chat.id])].value, reply_markup=key_for_test_1)
            elif call.data == 'yes_1':
                create_keyboard(call_create=call.message)
                schetchik_1[call.message.chat.id] += 1
            elif call.data == 'yes_2':
                schetchik_2[call.message.chat.id] += 1
                create_keyboard(call_create=call.message)
            elif call.data == 'yes_3':
                create_keyboard(call_create=call.message)
            elif call.data == 'no_1' or call.data == 'no_2' or call.data == 'no_3':
                create_keyboard(call_create=call.message)
            elif call.data == 'act_1' or call.data == 'act_2' or call.data == 'act_3':
                wb = openpyxl.reader.excel.load_workbook(filename="Ambulator.xlsx")
                wb.active = 0
                print("Программа зашла в акт")
                sheet = wb.active
                peremen = schetchik[call.message.chat.id]
                while sheet['E' + str(peremen)].value == None:
                    peremen -= 1
                bot.send_message(call.message.chat.id, sheet['E' + str(peremen)].value)
            elif call.data == 'comment_1' or call.data == 'comment_2' or call.data == 'comment_3':
                mesg = bot.send_message(call.message.chat.id, 'Введите комментарии')
                bot.register_next_step_handler(mesg, comment)
#----------------------------------------------НАЧИНАЮТСЯ СТАЦИОНАРНЫЕ УСЛОВИЯ------------------------------------------
            elif call.data == 'station_for_test':
                schetchik[call.message.chat.id] += 1
                key_for_test_1_st = types.InlineKeyboardMarkup()
                btn_yes_1_st = types.InlineKeyboardButton('Да', callback_data='yes_1_st')
                btn_no_1_st = types.InlineKeyboardButton('Нет', callback_data='no_1_st')
                btn_act_1_st = types.InlineKeyboardButton('Нормативный акт', callback_data='act_1_st')
                btn_comment_1_st = types.InlineKeyboardButton('Комментарии', callback_data='comment_1_st')
                key_for_test_1_st.row(btn_yes_1_st, btn_no_1_st)
                key_for_test_1_st.row(btn_act_1_st, btn_comment_1_st)
                wb = openpyxl.reader.excel.load_workbook(filename="Stationar.xlsx")
                wb.active = 0
                print("Программа зашла в статион фор тест который с клавиатуры")
                sheet = wb.active
                bot.send_message(call.message.chat.id, 'Вопрос №' + str(schetchik[call.message.chat.id]))
                bot.send_message(call.message.chat.id, sheet['A' + str(schetchik[call.message.chat.id])].value, reply_markup=key_for_test_1_st)
            elif call.data == 'yes_1_st':
                create_keyboard_st(call_create=call.message)
                schetchik_1[call.message.chat.id] += 1
            elif call.data == 'yes_2_st':
                create_keyboard_st(call_create=call.message)
                schetchik_2[call.message.chat.id] += 1
            elif call.data == 'yes_3_st':
                create_keyboard_st(call_create=call.message)
            elif call.data == 'no_1_st' or call.data == 'no_2_st' or call.data == 'no_3_st':
                create_keyboard_st(call_create=call.message)
            elif call.data == 'act_1_st' or call.data == 'act_2_st' or call.data == 'act_3_st':
                wb = openpyxl.reader.excel.load_workbook(filename="Stationar.xlsx")
                wb.active = 0
                print("Программа зашла в акт")
                sheet = wb.active
                peremen = schetchik[call.message.chat.id]
                while sheet['E' + str(peremen)].value == None:
                    peremen -= 1
                bot.send_message(call.message.chat.id, sheet['E' + str(peremen)].value)
            elif call.data == 'comment_1_st' or call.data == 'comment_2_st' or call.data == 'comment_3_st':
                mesg = bot.send_message(call.message.chat.id, 'Введите комментарии')
                bot.register_next_step_handler(mesg, comment)
    except Exception as e:
        print(repr(e))


bot.polling(none_stop=True)